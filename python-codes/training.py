#! C:/Users/WINDOWS/AppData/Local/Programs/Python/Python37/python.exe
#!/usr/bin/env python
# coding: utf-8

# In[12]:


import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.feature_selection import SelectFromModel
from sklearn.tree import export_graphviz
import matplotlib.pyplot as plt
import joblib as joblib
import openpyxl as xl
# import seaborn as sns

score_data = pd.read_excel('upload/demosaved (1).xlsx', engine='openpyxl')
# X = score_data.drop(columns=['RESULT'])
print(type(score_data) )
# y = score_data['RESULT']

# X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2)

# model = RandomForestClassifier(n_estimators = 100)

# model.fit(X_train, y_train)
# # print (y_test)
# predictions = model.predict(X_test)

# predictions = model.predict([
#   [0,0,0]
# ])
# score = accuracy_score(y_test, predictions)
# joblib.dump(model, 'student-assesstment.joblib')

# wb = xl.load_workbook('upload/haha.xlsx')
# sheet = wb['Sheet1']
# cell = sheet.cell(2, 4)
# cell.value = 'RESULT'
# print(cell.value)
# cell = sheet['a1']
# cell = sheet.cell(1,1)
# display value of the selected cell
# print(cell.value)

# for row in range(2, sheet.max_row + 1):
#   score1 = sheet.cell(row, 1)
#   score2 = sheet.cell(row, 2)
#   score3 = sheet.cell(row, 3)
  
 
#   model = joblib.load('student-assesstment.joblib')
#   predictions = model.predict([
#     [score1.value,score2.value,score3.value]
#   ])
#   # print(predictions)
#   resultCell = sheet.cell(row, 4)
#   resultCell.value = predictions[0]
# wb.save('upload/hehe.xlsx')


# print(predictions)
#print(score)

# Tree = model.estimators_[5]

# export_graphviz(
#   Tree,
#   out_file='student-assessment.dot',
#   feature_names=['Quiz1', 'Quiz2', 'Quiz3'],
#   class_names=sorted(y.unique()),
#   label='all',
#   rounded=True,
#   filled=True
# )
# import os
# os.environ["PATH"] += os.pathsep + 'D:/Program Files (x86)/Graphviz2.38/bin/'
# from graphviz import render
# render('dot', 'png', 'student-assessment.dot')


# %%
